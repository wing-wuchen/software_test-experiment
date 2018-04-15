from openpyxl import load_workbook
from selenium import webdriver

def read_from_xlsx(path):
    workbook = load_workbook(path)
    sheets = workbook.sheetnames  # 从名称获取sheet
    booksheet = workbook[sheets[0]]
    rows = booksheet.rows
    #迭代所有的行
    line = []
    for row in rows:
        line.append( [col.value for col in row] )
    #通过坐标读取值
    #cell_11 = booksheet.cell('a1').value
    #cell_11 = booksheet.cell(row=1, column=1).value
    return line

def main():
    pass

if __name__ == '__main__':
    main()
    book = read_from_xlsx("/users/wuchen/downloads/input.xlsx")
    driver = webdriver.Chrome()
    for row in book:
        s1,s2 = str(int(row[0])),str(row[1])
        driver.get('https://psych.liebes.top/st')
        driver.find_element_by_id("username").click()
        driver.find_element_by_id("username").clear()
        driver.find_element_by_id("username").send_keys(s1)
        driver.find_element_by_id("password").clear()
        driver.find_element_by_id("password").send_keys(s1[-6:])
        driver.find_element_by_id("submitButton").click()
        if driver.find_element_by_xpath("//p").text == s2:
            print(s1,'is ok')
        else:
            print(s1,"not ok")


        